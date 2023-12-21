from django.core.management.base import BaseCommand
import requests
import openpyxl

# from catalog.models import ExcelCategory
JWT = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ0b2tlbl90eXBlIjoiYWNjZXNzIiwiZXhwIjoxNzAzNTkzNDAwLCJpYXQiOjE3MDMxNjE0MDAsImp0aSI6ImNlOTFhMzM1OGU2NDQ0MGU5NTZiMjE2YzNjYmZmYWIwIiwidXNlcl9pZCI6MX0.jR0DfxuCyNVLxzco0YgMjAmh6EAXRKdcO8GWUALxHJY"
refresh = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ0b2tlbl90eXBlIjoicmVmcmVzaCIsImV4cCI6MTcwMzI0NzgwMCwiaWF0IjoxNzAzMTYxNDAwLCJqdGkiOiIxODg3NzM3ZGUxNWE0ZjE0YjY1ZTQzZDkzZjIzOTA5NiIsInVzZXJfaWQiOjF9.FqnADEz-nsRKEB-ra3LDUQ9l0C7Vnjn5TD6rZGDbj30"

url = 'http://127.0.0.1:8000/api/excel_category'


class Command(BaseCommand):
    help_text = 'Command to export data to xlxs file'

    def handle(self, *args, **options):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ['id', 'name', 'url', 'width', 'height']

        for num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=num, value=header)

        req = requests.get(url, headers={'Authorization': f'Bearer {JWT}'})
        # print(req)
        if req.status_code == 200:
            for row_num, obj in enumerate(req.json()['results'], 2):
                for col_num, item in enumerate(obj, 1):
                    sheet.cell(row=row_num, column=col_num, value=obj[item])

        workbook.save('excel_category.xlsx')
        print('Success!')
